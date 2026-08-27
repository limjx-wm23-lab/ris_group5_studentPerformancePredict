"""Student Performance Prediction System — polished final four-feature edition."""
from __future__ import annotations

from io import BytesIO
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sklearn.feature_selection import mutual_info_regression

from student_model import (
    CLASS_COLORS,
    CLASS_ORDER,
    FEATURES,
    FEATURE_LABELS,
    load_dataset,
    predict_batch,
    predict_model_details,
    student_recommendation,
    train_model_suite,
    validate_batch,
)

st.set_page_config(page_title="Student Performance Prediction", page_icon="🎓", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
:root{--navy:#0b1220;--blue:#2563eb;--violet:#7c3aed;--text:#172033;--muted:#64748b;--line:#dbe4f0}
.stApp{background:radial-gradient(circle at 18% 12%,rgba(96,165,250,.20),transparent 34%),radial-gradient(circle at 84% 84%,rgba(192,132,252,.16),transparent 32%),linear-gradient(180deg,#d6e4f7 0%,#e6edf8 48%,#eee9f8 100%);color:var(--text)}
.block-container{max-width:1100px;padding-top:.65rem;padding-bottom:2.2rem}
[data-testid="stSidebar"]{width:245px!important;min-width:245px!important;background:radial-gradient(circle at 14% 8%,rgba(99,102,241,.23),transparent 28%),linear-gradient(180deg,#08111f 0%,#111827 52%,#172554 100%);border-right:1px solid rgba(255,255,255,.08);box-shadow:10px 0 38px rgba(2,6,23,.15)}
[data-testid="stSidebar"]>div:first-child{width:245px!important;padding:1.15rem .95rem 1.6rem}[data-testid="stSidebar"] *{color:white}
[data-testid="stSidebar"] div[role="radiogroup"]{gap:.48rem}[data-testid="stSidebar"] div[role="radiogroup"] label{background:rgba(255,255,255,.045);border:1px solid rgba(255,255,255,.07);border-radius:14px;padding:.68rem .78rem;min-height:44px;transition:.18s ease}[data-testid="stSidebar"] div[role="radiogroup"] label:hover{background:rgba(255,255,255,.10);transform:translateX(3px)}[data-testid="stSidebar"] div[role="radiogroup"] label:has(input:checked){background:linear-gradient(90deg,#2563eb,#7c3aed);box-shadow:0 12px 28px rgba(79,70,229,.26)}
.sidebar-brand{font-size:1.55rem;font-weight:900;letter-spacing:-.03em}.sidebar-sub{font-size:.72rem;color:#afc6ff;margin:.15rem 0 1.1rem}.sidebar-foot{margin-top:1.35rem;padding-top:.95rem;border-top:1px solid rgba(255,255,255,.12);color:#b8c7e8;font-size:.70rem;line-height:1.65}.sidebar-pill{display:inline-block;margin-top:.5rem;padding:.23rem .5rem;border-radius:999px;background:rgba(37,99,235,.18);border:1px solid rgba(147,197,253,.22);color:#dbeafe;font-weight:800}
.hero{padding:1.42rem 1.55rem;border-radius:27px;color:white;background:radial-gradient(circle at 92% 10%,rgba(124,58,237,.34),transparent 30%),linear-gradient(135deg,#0b1220 0%,#172554 58%,#4c1d95 100%);box-shadow:0 22px 52px rgba(15,23,42,.23);margin-bottom:1rem;border:1px solid rgba(255,255,255,.09)}
.hero .eyebrow{font-size:.72rem;font-weight:900;letter-spacing:.14em;text-transform:uppercase;color:#bfdbfe}.hero h1{margin:.28rem 0 .45rem;color:white;font-size:2rem;letter-spacing:-.04em}.hero p{margin:0;color:#dbeafe;font-size:.93rem;line-height:1.55;max-width:880px}
[data-testid="stMetric"],[data-testid="stForm"],.card,.feature-card,.mode-card,.why-card{background:linear-gradient(145deg,rgba(255,255,255,.94),rgba(246,249,255,.82));border:1px solid rgba(148,163,184,.26);border-radius:20px;box-shadow:0 15px 38px rgba(37,99,235,.08),inset 0 1px 0 rgba(255,255,255,.92);backdrop-filter:blur(14px)}
[data-testid="stMetric"]{padding:.76rem .86rem}[data-testid="stMetricLabel"] p{color:var(--muted)!important;font-weight:750!important;font-size:.80rem!important}[data-testid="stMetricValue"]>div{font-size:1.42rem!important;font-weight:900!important;color:var(--text)!important}[data-testid="stForm"]{padding:1rem 1.05rem 1.1rem}
.feature-card{padding:1rem;min-height:176px;transition:.2s ease}.feature-card:hover,.mode-card:hover{transform:translateY(-4px);box-shadow:0 22px 46px rgba(79,70,229,.13)}.feature-number{width:38px;height:38px;border-radius:12px;display:flex;align-items:center;justify-content:center;color:white;font-weight:900;background:linear-gradient(135deg,#2563eb,#7c3aed);margin-bottom:.72rem}.feature-title{font-size:1.02rem;font-weight:900;color:#172033;margin-bottom:.35rem}.feature-desc{font-size:.81rem;color:#64748b;line-height:1.52}.feature-tag{display:inline-block;margin-top:.62rem;padding:.22rem .47rem;border-radius:999px;background:#eef2ff;color:#4338ca;font-size:.69rem;font-weight:850}
.hub{background:linear-gradient(145deg,rgba(248,251,255,.86),rgba(244,245,255,.80));border:1px solid rgba(148,163,184,.24);border-radius:28px;padding:1.15rem;box-shadow:0 18px 48px rgba(37,99,235,.08)}.hub-welcome{background:linear-gradient(135deg,#0f172a 0%,#1e3a8a 58%,#5b21b6 100%);border-radius:24px;padding:1.1rem 1.3rem;color:white;margin-bottom:1rem;box-shadow:0 20px 50px rgba(37,99,235,.20)}.hub-welcome b{font-size:1.35rem}.hub-welcome p{margin:.35rem 0 0;color:#e2e8f0;font-size:.84rem}.mode-card{padding:.9rem .95rem .78rem;min-height:178px;margin-bottom:.55rem}.mode-icon{width:50px;height:50px;border-radius:15px;display:flex;align-items:center;justify-content:center;font-size:1.5rem;background:linear-gradient(135deg,#dbeafe,#ede9fe);margin-bottom:.7rem}.mode-title{font-size:1.13rem;font-weight:900;color:#172033;margin-bottom:.34rem}.mode-desc{font-size:.81rem;line-height:1.45;color:#64748b;min-height:48px}.chips{display:flex;gap:.35rem;flex-wrap:wrap;margin-top:.58rem}.chip{padding:.24rem .45rem;border-radius:999px;background:#f8fafc;border:1px solid #e2e8f0;color:#334155;font-size:.68rem;font-weight:800}.why-card{padding:1rem;margin-top:.9rem}.why-grid{display:grid;grid-template-columns:repeat(5,1fr);gap:.55rem}.why-item{background:linear-gradient(135deg,#eff6ff,#f5f3ff);border:1px solid #dbeafe;border-radius:13px;padding:.68rem .55rem;text-align:center;font-size:.75rem;font-weight:800;color:#334155}
.mode-hero{background:linear-gradient(120deg,#dcecff,#e8edff 55%,#efe7ff);border:1px solid #dbeafe;border-radius:20px;padding:.88rem 1rem;margin-bottom:.8rem;box-shadow:0 10px 26px rgba(37,99,235,.08)}.mode-hero small{font-size:.68rem;font-weight:900;letter-spacing:.12em;text-transform:uppercase;color:#4338ca}.mode-hero h2{margin:.22rem 0 .30rem!important;font-size:1.26rem!important;color:#172033!important}.mode-hero p{margin:0;color:#475569;font-size:.81rem;line-height:1.48}
.stButton>button,[data-testid="stFormSubmitButton"]>button{width:100%;border:none;border-radius:14px;font-weight:850;color:white;min-height:45px;background:linear-gradient(90deg,#2563eb,#4f46e5 52%,#7c3aed);box-shadow:0 11px 25px rgba(79,70,229,.20)}.stDownloadButton>button{width:100%;border-radius:14px;font-weight:850;color:white;min-height:45px;background:linear-gradient(135deg,#10b981,#059669);border:1px solid rgba(255,255,255,.18)}
.result-card{border-radius:22px;padding:1.05rem 1.2rem;background:linear-gradient(135deg,#0f172a,#1e3a8a 62%,#4c1d95);color:white;box-shadow:0 18px 42px rgba(30,58,138,.20);margin:.7rem 0}.result-card .label{font-size:.75rem;color:#bfdbfe;font-weight:800;text-transform:uppercase;letter-spacing:.08em}.result-card .big{font-size:2rem;font-weight:950;letter-spacing:-.035em;margin:.15rem 0}.result-card .small{font-size:.82rem;color:#dbeafe}.cgpa-grid{display:grid;grid-template-columns:repeat(4,1fr);gap:.55rem;margin:.65rem 0 .85rem}.cgpa-box{background:rgba(255,255,255,.93);border:1px solid #e2e8f0;border-radius:14px;padding:.65rem .70rem}.cgpa-box b{display:block;font-size:.82rem;color:#172033}.cgpa-box span{font-size:.72rem;color:#64748b}.info-box{background:linear-gradient(135deg,#eff6ff,#eef2ff);border:1px solid #dbeafe;border-radius:15px;padding:.70rem .82rem;color:#334155;font-size:.80rem;line-height:1.48;margin:.55rem 0 .8rem}
[data-testid="stDataFrame"],[data-testid="stPlotlyChart"]{background:rgba(255,255,255,.94);border:1px solid rgba(148,163,184,.22);border-radius:17px;overflow:hidden;box-shadow:0 13px 32px rgba(37,99,235,.07)}
@media(max-width:900px){.why-grid{grid-template-columns:repeat(2,1fr)}.cgpa-grid{grid-template-columns:repeat(2,1fr)}.hero h1{font-size:1.55rem}}
</style>
""", unsafe_allow_html=True)

@st.cache_data(show_spinner=False)
def get_dataset():
    return load_dataset()

@st.cache_resource(show_spinner="Training KNN, SVM and ANN...")
def get_suite():
    data, _ = load_dataset()
    return train_model_suite(data)

@st.cache_data(show_spinner=False)
def feature_evidence(data: pd.DataFrame) -> pd.DataFrame:
    X = data[FEATURES].apply(pd.to_numeric)
    y = pd.to_numeric(data["Final_CGPA"])
    mi = mutual_info_regression(X, y, random_state=42)
    rows=[]
    for i,f in enumerate(FEATURES):
        rows.append({"Feature":FEATURE_LABELS[f],"Mutual Information":mi[i],"Pearson Correlation":X[f].corr(y,method="pearson"),"Spearman Correlation":X[f].corr(y,method="spearman")})
    return pd.DataFrame(rows).sort_values("Pearson Correlation",ascending=False).reset_index(drop=True)

def hero(title, subtitle, eyebrow="RIS Group 5 • BMCS2003 Artificial Intelligence"):
    st.markdown(f'<div class="hero"><div class="eyebrow">{eyebrow}</div><h1>{title}</h1><p>{subtitle}</p></div>',unsafe_allow_html=True)

def mode_hero(title, subtitle, eyebrow):
    st.markdown(f'<div class="mode-hero"><small>{eyebrow}</small><h2>{title}</h2><p>{subtitle}</p></div>',unsafe_allow_html=True)

def cgpa_guide():
    st.markdown('''<div class="cgpa-grid"><div class="cgpa-box"><b>🌟 Excellent</b><span>Final CGPA 3.50 – 4.00</span></div><div class="cgpa-box"><b>✅ Good</b><span>Final CGPA 3.00 – 3.49</span></div><div class="cgpa-box"><b>📘 Average</b><span>Final CGPA 2.50 – 2.99</span></div><div class="cgpa-box"><b>⚠️ At Risk</b><span>Final CGPA below 2.50</span></div></div>''',unsafe_allow_html=True)

def four_feature_note():
    st.markdown('<div class="info-box"><b>Exactly 4 ML inputs:</b> Average Score, Attendance Percentage, Study Hours per Day and Previous CGPA. The same four values are used by KNN, SVM and ANN.</div>',unsafe_allow_html=True)

def feature_cards():
    descriptions={"Average_Score":"Current assessment performance and a strong academic signal.","Attendance_Pct":"Participation and consistency in scheduled learning activities.","Study_Hours_Per_Day":"Daily academic effort, preparation and independent study.","Previous_CGPA":"Historical academic achievement and strongest selected predictor."}
    tags={"Average_Score":"Academic performance","Attendance_Pct":"Engagement","Study_Hours_Per_Day":"Learning effort","Previous_CGPA":"Academic history"}
    cols=st.columns(4)
    for i,f in enumerate(FEATURES,1):
        cols[i-1].markdown(f'<div class="feature-card"><div class="feature-number">{i}</div><div class="feature-title">{FEATURE_LABELS[f]}</div><div class="feature-desc">{descriptions[f]}</div><div class="feature-tag">{tags[f]}</div></div>',unsafe_allow_html=True)

def autosize(ws):
    for cells in ws.columns:
        ws.column_dimensions[get_column_letter(cells[0].column)].width=min(max(max(len(str(c.value or "")) for c in cells)+3,12),36)

def excel_bytes(dataframe,title):
    wb=Workbook(); ws=wb.active; ws.title="Prediction Results"
    dark=PatternFill("solid",fgColor="172554"); purple=PatternFill("solid",fgColor="6D28D9"); alt=PatternFill("solid",fgColor="F8FAFC"); thin=Side(style="thin",color="D9E2F1")
    ws.merge_cells(start_row=1,start_column=1,end_row=1,end_column=max(1,len(dataframe.columns)))
    c=ws.cell(1,1,title); c.fill=dark; c.font=Font(color="FFFFFF",bold=True,size=16); c.alignment=Alignment(horizontal="center")
    for j,col in enumerate(dataframe.columns,1):
        c=ws.cell(3,j,col); c.fill=purple if col=="Final_Prediction" else dark; c.font=Font(color="FFFFFF",bold=True); c.alignment=Alignment(horizontal="center",wrap_text=True)
    for i,row in enumerate(dataframe.itertuples(index=False),4):
        for j,value in enumerate(row,1):
            c=ws.cell(i,j,value); c.fill=alt if i%2==0 else PatternFill("solid",fgColor="FFFFFF"); c.border=Border(left=thin,right=thin,top=thin,bottom=thin); c.alignment=Alignment(horizontal="center")
            if "Confidence" in str(dataframe.columns[j-1]) and isinstance(value,(float,int)): c.number_format="0.0%"
    ws.freeze_panes="A4"; ws.auto_filter.ref=ws.dimensions; ws.sheet_view.showGridLines=False; autosize(ws)
    out=BytesIO(); wb.save(out); return out.getvalue()

def batch_template():
    sample=pd.DataFrame([{"Student_ID":"S001","Student_Name":"Ali Tan","Average_Score":82.0,"Attendance_Pct":91.0,"Study_Hours_Per_Day":4.5,"Previous_CGPA":3.45},{"Student_ID":"S002","Student_Name":"Mei Ling","Average_Score":67.0,"Attendance_Pct":78.0,"Study_Hours_Per_Day":2.5,"Previous_CGPA":2.85}])
    return excel_bytes(sample,"Batch Prediction Template — Exactly 4 ML Features")

def confusion_fig(matrix,classes,name):
    fig=go.Figure(data=go.Heatmap(z=matrix,x=classes,y=classes,colorscale="Blues",text=matrix,texttemplate="%{text}")); fig.update_layout(title=dict(text=f"{name} Confusion Matrix",x=.5,xanchor="center"),xaxis_title="Predicted Class",yaxis_title="Actual Class",height=420); return fig

df,data_source=get_dataset(); suite=get_suite(); evaluation=suite.evaluation.copy(); best_model=suite.best_model; best_accuracy=float(evaluation.iloc[0]["Accuracy"]); evidence=feature_evidence(df)

st.sidebar.markdown('<div class="sidebar-brand">🎓 Student AI</div><div class="sidebar-sub">Student Performance Prediction</div>',unsafe_allow_html=True)
st.sidebar.markdown("### Navigation")
choice=st.sidebar.radio("Navigation Menu",["🏠 Home","🎯 Prediction","📊 Model Results","🔗 Correlation","📈 Dataset","⭐ Feature Analysis","ℹ️ About"],label_visibility="collapsed")
page=choice.split(" ",1)[1]
st.sidebar.markdown(f'<div class="sidebar-foot"><b>Final Model Inputs</b><br>1. Average Score<br>2. Attendance Percentage<br>3. Study Hours per Day<br>4. Previous CGPA<br><span class="sidebar-pill">Exactly 4 ML Features</span><br><br><b>Best Model:</b> {best_model}<br><b>Accuracy:</b> {best_accuracy:.1%}<br><br>Developed by RIS Group 5<br>© 2026 All Rights Reserved</div>',unsafe_allow_html=True)
if page!="Prediction": st.session_state.pop("prediction_mode",None)

if page=="Home":
    hero("Student Performance Prediction System","AI-powered academic performance analysis using KNN, SVM and ANN with exactly four selected model inputs.")
    c1,c2,c3,c4=st.columns(4); c1.metric("Students",f"{len(df):,}"); c2.metric("ML Features","4"); c3.metric("Best Model",best_model); c4.metric("Best Accuracy",f"{best_accuracy:.1%}")
    st.markdown("### Four Final Prediction Features"); feature_cards()
    long=evaluation.melt(id_vars="Model",value_vars=["Accuracy","Precision","Recall","F1 Score"],var_name="Metric",value_name="Score"); fig=px.bar(long,x="Model",y="Score",color="Metric",barmode="group",text_auto=".1%",range_y=[0,1],title="Model Performance Overview"); fig.update_yaxes(tickformat=".0%"); fig.update_layout(title=dict(x=.5,xanchor="center"),height=420); st.plotly_chart(fig,use_container_width=True)
    cgpa_guide(); four_feature_note()

elif page=="Prediction" and not st.session_state.get("prediction_mode"):
    hero("Prediction Hub","Choose how you want to analyse student performance using the final four-feature KNN, SVM and ANN system.","AI-Powered Academic Intelligence")
    st.markdown('<div class="hub">',unsafe_allow_html=True); st.markdown('<div class="hub-welcome"><b>Choose Your Prediction Mode</b><p>Analyse one student in detail or process many students through the same validated four-feature pipeline.</p></div>',unsafe_allow_html=True)
    a,b,c,d=st.columns(4); a.metric("🏆 Best Model",best_model); b.metric("🎯 Best Accuracy",f"{best_accuracy:.1%}"); c.metric("📚 Input Features","4"); d.metric("🤖 ML Models","3")
    l,r=st.columns(2)
    with l:
        st.markdown('<div class="mode-card"><div class="mode-icon">👤</div><div class="mode-title">Individual Prediction</div><div class="mode-desc">Enter one student’s four academic inputs and compare KNN, SVM and ANN predictions in real time.</div><div class="chips"><div class="chip">✓ 4 inputs only</div><div class="chip">✓ 3-model comparison</div><div class="chip">✓ Excel report</div></div></div>',unsafe_allow_html=True)
        if st.button("🚀 Start Individual Prediction",key="start_individual",type="primary"): st.session_state["prediction_mode"]="individual"; st.rerun()
    with r:
        st.markdown('<div class="mode-card"><div class="mode-icon">📂</div><div class="mode-title">Batch Prediction</div><div class="mode-desc">Upload Excel or CSV records containing the same four required ML inputs for multiple students.</div><div class="chips"><div class="chip">✓ Excel / CSV</div><div class="chip">✓ 4 required inputs</div><div class="chip">✓ At-risk detection</div></div></div>',unsafe_allow_html=True)
        if st.button("📂 Start Batch Prediction",key="start_batch",type="primary"): st.session_state["prediction_mode"]="batch"; st.rerun()
    st.markdown('<div class="why-card"><b>Why Use This System?</b><div class="why-grid"><div class="why-item">AI-Powered Prediction</div><div class="why-item">Exactly 4 Inputs</div><div class="why-item">Batch Processing</div><div class="why-item">Professional Excel Reports</div><div class="why-item">Early At-Risk Identification</div></div></div></div>',unsafe_allow_html=True)

elif page=="Prediction" and st.session_state.get("prediction_mode")=="individual":
    mode_hero("👤 Individual Prediction","Only the four numbered academic fields below are passed into the machine-learning models.","Single-Student Analysis • 4 ML Inputs")
    if st.button("← Back to Prediction Hub",key="back_individual"): st.session_state.pop("prediction_mode",None); st.session_state.pop("individual_result",None); st.rerun()
    cgpa_guide()
    with st.form("individual_form"):
        i1,i2=st.columns(2); student_id=i1.text_input("Student ID (optional)",placeholder="e.g. S001"); student_name=i2.text_input("Student Name (optional)",placeholder="e.g. Alex Tan")
        st.markdown("### Four Prediction Features"); l,r=st.columns(2); average=l.number_input("1. Average Score (%)",0.0,100.0,75.0,1.0); attendance=r.number_input("2. Attendance Percentage (%)",0.0,100.0,85.0,.5); study=l.number_input("3. Study Hours per Day",0.0,24.0,3.0,.5); previous=r.number_input("4. Previous CGPA",0.0,4.0,3.0,.01,format="%.2f"); submitted=st.form_submit_button("🚀 Predict Student Performance",type="primary",use_container_width=True)
    if submitted:
        X=pd.DataFrame([{"Average_Score":average,"Attendance_Pct":attendance,"Study_Hours_Per_Day":study,"Previous_CGPA":previous}])[FEATURES]; comp=predict_model_details(X,suite); best=comp[comp["Model"]==best_model].iloc[0]; st.session_state["individual_result"]={"comp":comp,"pred":str(best["Prediction"]),"conf":float(best["Confidence"]),"id":student_id,"name":student_name,"inputs":X.iloc[0].to_dict()}
    saved=st.session_state.get("individual_result")
    if saved:
        emoji={"Excellent":"🌟","Good":"✅","Average":"📘","At Risk":"⚠️"}.get(saved["pred"],"🎓"); st.markdown(f'<div class="result-card"><div class="label">Final Prediction • {best_model}</div><div class="big">{emoji} {saved["pred"]}</div><div class="small">Model confidence: {saved["conf"]:.1%}</div></div>',unsafe_allow_html=True); st.info(student_recommendation(saved["pred"])); show=saved["comp"].copy(); show["Confidence"]=show["Confidence"].map(lambda x:f"{x:.1%}"); st.dataframe(show,hide_index=True,use_container_width=True)
        report=pd.DataFrame([{"Student_ID":saved["id"] or "N/A","Student_Name":saved["name"] or "N/A",**saved["inputs"],"Final_Prediction":saved["pred"],"Final_Confidence":saved["conf"],"Best_Model":best_model}]); st.download_button("⬇️ Download Prediction Report",excel_bytes(report,"Individual Prediction Report — 4 ML Features"),"individual_prediction.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

elif page=="Prediction" and st.session_state.get("prediction_mode")=="batch":
    mode_hero("📂 Batch Prediction","Student ID and Student Name are optional. Exactly four academic input columns are required.","Multi-Student Analysis • 4 Required ML Inputs")
    if st.button("← Back to Prediction Hub",key="back_batch"): st.session_state.pop("prediction_mode",None); st.session_state.pop("batch_result",None); st.rerun()
    d1,d2=st.columns([1,2]); d1.download_button("⬇️ Download Excel Template",batch_template(),"student_batch_prediction_template.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True); d2.markdown('<div class="info-box"><b>Exactly 4 required ML columns:</b> Average_Score, Attendance_Pct, Study_Hours_Per_Day and Previous_CGPA.<br>Student_ID and Student_Name are optional.</div>',unsafe_allow_html=True)
    up=st.file_uploader("Upload completed Excel or CSV file",type=["xlsx","xls","csv"])
    if up is not None:
        try: batch_df=pd.read_csv(up) if up.name.lower().endswith(".csv") else pd.read_excel(up)
        except Exception as exc: st.error(f"Unable to read file: {exc}")
        else:
            errors=validate_batch(batch_df)
            if errors:
                st.error("Please correct the following validation problem(s):"); [st.write(f"• {e}") for e in errors]
            else:
                x1,x2,x3=st.columns(3); x1.metric("Uploaded Students",f"{len(batch_df):,}"); x2.metric("Required Features","4"); x3.metric("Final Model",best_model); preview=[c for c in ["Student_ID","Student_Name",*FEATURES] if c in batch_df.columns]; st.dataframe(batch_df[preview].head(50),hide_index=True,use_container_width=True)
                if st.button("🚀 Run Batch Prediction",key="run_batch",type="primary"): st.session_state["batch_result"]=predict_batch(batch_df,suite)
    result=st.session_state.get("batch_result")
    if isinstance(result,pd.DataFrame):
        counts=result["Final_Prediction"].value_counts(); c1,c2,c3,c4,c5=st.columns(5); c1.metric("Total",len(result)); c2.metric("Excellent",int(counts.get("Excellent",0))); c3.metric("Good",int(counts.get("Good",0))); c4.metric("Average",int(counts.get("Average",0))); c5.metric("At Risk",int(counts.get("At Risk",0)))
        cols=[c for c in ["Student_ID","Student_Name",*FEATURES,"Final_Prediction","Final_Confidence"] if c in result.columns]; show=result[cols].copy(); show["Final_Confidence"]=show["Final_Confidence"].map(lambda x:f"{x:.1%}"); st.dataframe(show,hide_index=True,use_container_width=True,height=420)
        export=[c for c in ["Student_ID","Student_Name",*FEATURES,"KNN_Prediction","SVM_Prediction","ANN_Prediction","Final_Prediction","Final_Confidence"] if c in result.columns]; st.download_button("⬇️ Download Complete Batch Report",excel_bytes(result[export],"Batch Prediction Report — 4 ML Features"),"batch_prediction_results.xlsx","application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)

elif page=="Model Results":
    hero("Model Results","Compare KNN, SVM and ANN using the same four input features and the same 80/20 stratified hold-out split.","Machine Learning Evaluation")
    a,b,c,d=st.columns(4); a.metric("Best Model",best_model); b.metric("Best Accuracy",f"{best_accuracy:.1%}"); c.metric("Training Records",f"{suite.train_size:,}"); d.metric("Testing Records",f"{suite.test_size:,}")
    table=evaluation.copy(); [table.__setitem__(col,table[col].map(lambda x:f"{x:.2%}")) for col in ["Accuracy","Precision","Recall","F1 Score"]]; st.dataframe(table,hide_index=True,use_container_width=True)
    long=evaluation.melt(id_vars="Model",value_vars=["Accuracy","Precision","Recall","F1 Score"],var_name="Metric",value_name="Score"); fig=px.bar(long,x="Metric",y="Score",color="Model",barmode="group",text_auto=".1%",range_y=[0,1],title="KNN vs SVM vs ANN"); fig.update_yaxes(tickformat=".0%"); fig.update_layout(title=dict(x=.5,xanchor="center")); st.plotly_chart(fig,use_container_width=True)
    tabs=st.tabs(list(evaluation["Model"]));
    for tab,name in zip(tabs,list(evaluation["Model"])):
        with tab: st.plotly_chart(confusion_fig(suite.confusion_matrices[name],suite.classes,name),use_container_width=True)
    four_feature_note()

elif page=="Correlation":
    hero("Feature Correlation Analysis","Examine only the relationships between the final four ML inputs and Final CGPA.","Evidence-Based Feature Selection • Exactly 4 Features")
    a,b,c=st.columns(3); a.metric("Students",f"{len(df):,}"); b.metric("Selected Features","4"); c.metric("Target","Final CGPA")
    corr=df[FEATURES+["Final_CGPA"]].corr().round(3); labels={**FEATURE_LABELS,"Final_CGPA":"Final CGPA"}; display=corr.rename(index=labels,columns=labels); fig=px.imshow(display,text_auto=".2f",color_continuous_scale="RdBu_r",zmin=-1,zmax=1,aspect="auto",title="Four-Feature Correlation Matrix"); fig.update_layout(title=dict(x=.5,xanchor="center"),height=520); st.plotly_chart(fig,use_container_width=True)
    table=evidence[["Feature","Pearson Correlation","Spearman Correlation"]].copy().round(4); st.dataframe(table,hide_index=True,use_container_width=True); rank=table.sort_values("Pearson Correlation"); fig=px.bar(rank,x="Pearson Correlation",y="Feature",orientation="h",text=rank["Pearson Correlation"].map(lambda v:f"{v:+.3f}"),range_x=[0,1],title="Correlation Strength of the Four Selected Features"); fig.update_traces(textposition="outside"); fig.update_layout(title=dict(x=.5,xanchor="center")); st.plotly_chart(fig,use_container_width=True); four_feature_note()

elif page=="Dataset":
    hero("Dataset Explorer","Explore the 5,000 student records. The default table highlights the four ML inputs plus identifier and target information.","University Student Performance Dataset")
    a,b,c,d=st.columns(4); a.metric("Student Records",f"{len(df):,}"); b.metric("ML Inputs","4"); c.metric("Missing in ML Inputs",int(df[FEATURES].isna().sum().sum())); d.metric("Target Classes","4")
    default=[c for c in ["Student_ID",*FEATURES,"Final_CGPA","Performance_Category"] if c in df.columns]; full=st.toggle("Show all source dataset columns",value=False); view=df.copy() if full else df[default].copy()
    if full: st.info("Additional source columns are shown only for dataset exploration. The prediction models still use exactly four ML inputs.")
    search=st.text_input("Search Student ID",placeholder="e.g. ID00001");
    if search and "Student_ID" in view.columns: view=view[view["Student_ID"].astype(str).str.contains(search,case=False,na=False)]
    cats=st.multiselect("Performance Category",CLASS_ORDER,default=CLASS_ORDER)
    if "Performance_Category" in view.columns and cats: view=view[view["Performance_Category"].astype(str).isin(cats)]
    st.dataframe(view,hide_index=True,use_container_width=True,height=500)
    counts=df["Performance_Category"].astype(str).value_counts().reindex(CLASS_ORDER,fill_value=0).reset_index(); counts.columns=["Performance Category","Students"]; fig=px.pie(counts,names="Performance Category",values="Students",color="Performance Category",color_discrete_map=CLASS_COLORS,hole=.42,title="Performance Category Distribution"); fig.update_layout(title=dict(x=.5,xanchor="center")); st.plotly_chart(fig,use_container_width=True)

elif page=="Feature Analysis":
    hero("Feature Analysis","Evidence-based analysis of the exact four features used by KNN, SVM and ANN. The table below is intentionally limited to four rows.","Final Feature Set • 4 Features Only")
    a,b,c,d=st.columns(4); a.metric("Final Features","4"); b.metric("Strongest Correlation",evidence.iloc[0]["Feature"]); c.metric("Best Pearson r",f"{evidence.iloc[0]['Pearson Correlation']:.3f}"); d.metric("Models Using Set","3")
    st.markdown("### Final Four Features"); feature_cards(); st.markdown("### Feature Selection Evidence"); show=evidence.copy().round(4); st.dataframe(show,hide_index=True,use_container_width=True)
    chart=show.sort_values("Pearson Correlation"); fig=px.bar(chart,x="Pearson Correlation",y="Feature",orientation="h",text=chart["Pearson Correlation"].map(lambda v:f"{v:.3f}"),range_x=[0,1],title="Strength of Relationship with Final CGPA"); fig.update_traces(textposition="outside"); fig.update_layout(title=dict(x=.5,xanchor="center")); st.plotly_chart(fig,use_container_width=True); four_feature_note()

elif page=="About":
    hero("About the Project","A complete student-performance classification prototype developed for BMCS2003 Artificial Intelligence using Streamlit and Scikit-learn.","Project Documentation")
    a,b,c=st.columns(3); a.markdown('<div class="card" style="padding:1rem"><b>🎯 Objective</b><p>Predict performance category and support early academic intervention through individual and batch workflows.</p></div>',unsafe_allow_html=True); b.markdown('<div class="card" style="padding:1rem"><b>🧠 Models</b><p>KNN<br>SVM<br>ANN<br>80/20 stratified hold-out evaluation</p></div>',unsafe_allow_html=True); c.markdown('<div class="card" style="padding:1rem"><b>📚 Exactly 4 ML Inputs</b><p>Average Score<br>Attendance Percentage<br>Study Hours per Day<br>Previous CGPA</p></div>',unsafe_allow_html=True)
    st.markdown("### Target Classes"); cgpa_guide(); st.markdown("### Responsible Use"); st.warning("This is an academic decision-support prototype. Predictions should not be the sole basis for high-impact education decisions. Human review remains necessary."); four_feature_note()
